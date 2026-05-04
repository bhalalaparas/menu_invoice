"""
Logging and XLSX audit trail for every Bedrock invoke (menu + invoice).
"""

from __future__ import annotations

import logging
import os
import threading
from datetime import datetime, timezone
from typing import Any, Dict, Optional, Tuple

from openpyxl import Workbook, load_workbook

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
OUTPUT_DIR = os.path.join(BASE_DIR, "output")
os.makedirs(OUTPUT_DIR, exist_ok=True)

BEDROCK_USAGE_XLSX = os.path.join(OUTPUT_DIR, "bedrock_usage_log.xlsx")
MAX_CELL_CHARS = 32000

_XLSX_LOCK = threading.Lock()

HEADERS = [
    "timestamp_utc",
    "source",
    "model_id",
    "input_tokens",
    "output_tokens",
    "total_tokens",
    "prompt",
    "output",
]


def setup_bedrock_logger() -> logging.Logger:
    log = logging.getLogger("menu_po.bedrock")
    if log.handlers:
        return log
    log.setLevel(logging.INFO)
    path = os.path.join(OUTPUT_DIR, "bedrock.log")
    fh = logging.FileHandler(path, encoding="utf-8")
    fh.setFormatter(
        logging.Formatter("%(asctime)s | %(levelname)s | %(message)s")
    )
    log.addHandler(fh)
    log.propagate = False
    return log


logger = setup_bedrock_logger()


def _truncate(s: str) -> str:
    if not s:
        return ""
    if len(s) <= MAX_CELL_CHARS:
        return s
    return s[: MAX_CELL_CHARS - 20] + "\n...[truncated]"


def extract_usage_from_response(parsed: Dict[str, Any]) -> Tuple[Optional[int], Optional[int]]:
    usage = parsed.get("usage") or {}
    if not isinstance(usage, dict):
        return None, None
    inp = usage.get("input_tokens")
    out = usage.get("output_tokens")
    try:
        inp_i = int(inp) if inp is not None else None
    except (TypeError, ValueError):
        inp_i = None
    try:
        out_i = int(out) if out is not None else None
    except (TypeError, ValueError):
        out_i = None
    return inp_i, out_i


def append_bedrock_usage_row(
    *,
    source: str,
    model_id: str,
    prompt: str,
    output: str,
    input_tokens: Optional[int],
    output_tokens: Optional[int],
) -> None:
    total = None
    if input_tokens is not None and output_tokens is not None:
        total = input_tokens + output_tokens
    row = [
        datetime.now(timezone.utc).isoformat(),
        source,
        model_id,
        input_tokens if input_tokens is not None else "",
        output_tokens if output_tokens is not None else "",
        total if total is not None else "",
        _truncate(prompt),
        _truncate(output),
    ]
    with _XLSX_LOCK:
        if not os.path.isfile(BEDROCK_USAGE_XLSX):
            wb = Workbook()
            ws = wb.active
            assert ws is not None
            ws.title = "bedrock_calls"
            ws.append(HEADERS)
            wb.save(BEDROCK_USAGE_XLSX)
        wb = load_workbook(BEDROCK_USAGE_XLSX)
        ws = wb.active
        assert ws is not None
        ws.append(row)
        wb.save(BEDROCK_USAGE_XLSX)

    logger.info(
        "source=%s model=%s in_tok=%s out_tok=%s",
        source,
        model_id,
        input_tokens,
        output_tokens,
    )


def log_bedrock_call(
    *,
    source: str,
    model_id: str,
    prompt: str,
    output: str,
    parsed_response: Dict[str, Any],
) -> None:
    inp, out = extract_usage_from_response(parsed_response)
    append_bedrock_usage_row(
        source=source,
        model_id=model_id,
        prompt=prompt,
        output=output,
        input_tokens=inp,
        output_tokens=out,
    )
